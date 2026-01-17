"""
Data loader module for Budget Planning App
Extracts and processes data from the Excel budget model
"""
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import numpy as np

class BudgetDataLoader:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.territories = ['UK', 'ES', 'DE', 'IT', 'FR', 'RO', 'CZ', 'HU', 'SK', 'Other EU', 'ROW']
        self.territory_groups = {
            'UK': 'UK',
            'ES': 'CE', 'DE': 'CE', 'IT': 'CE', 'FR': 'CE',
            'RO': 'EE', 'CZ': 'EE', 'HU': 'EE', 'SK': 'EE',
            'Other EU': 'CE', 'ROW': 'ROW'
        }
        self.channels = ['DTC', 'B2B', 'Marketplace', 'TikTok']
        # FY27 period: 2026-02 to 2027-01
        self.fy27_start = '2026-02'
        self.fy27_end = '2027-01'

    def filter_fy27_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Filter dataframe to only include FY27 date columns (2026-02 to 2027-01)"""
        date_cols = [c for c in df.columns if isinstance(c, str) and c.startswith('20')]
        fy27_cols = [c for c in date_cols if self.fy27_start <= c <= self.fy27_end]
        non_date_cols = [c for c in df.columns if c not in date_cols]
        return df[non_date_cols + fy27_cols]

    def load_b2b_data(self) -> pd.DataFrame:
        """Load B2B customer revenue data"""
        df = pd.read_excel(self.file_path, sheet_name='B2B', header=5)

        # Clean column names - convert datetime to string
        new_cols = []
        for col in df.columns:
            if isinstance(col, datetime):
                new_cols.append(col.strftime('%Y-%m'))
            else:
                new_cols.append(str(col))
        df.columns = new_cols

        # Map alternative column names to standard names
        column_mapping = {}
        for col in df.columns:
            col_lower = col.lower().strip()
            if 'margin' in col_lower and 'customer' in col_lower:
                column_mapping[col] = 'Customer Margin'
            elif col_lower == 'margin':
                column_mapping[col] = 'Customer Margin'

        if column_mapping:
            df = df.rename(columns=column_mapping)

        # Keep only relevant columns (use only those that exist)
        base_cols = ['Customer Name', 'Country', 'Country Group', 'Customer Margin', 'Last Year FY25 Revenue']
        date_cols = [c for c in df.columns if c.startswith('202')]

        # Filter to only existing base columns
        existing_base_cols = [c for c in base_cols if c in df.columns]
        df = df[existing_base_cols + date_cols].copy()

        # Clean data
        df = df.dropna(subset=['Customer Name'])

        # Filter to actual customer revenue rows only:
        # 1. Revenue rows have Customer Margin = NaN (cost rows have string values like "Retros")
        # 2. Actual customers have valid Country AND Country Group (aggregation rows have NaN)
        if 'Customer Margin' in df.columns:
            # Keep only revenue rows (Customer Margin is NaN)
            df = df[df['Customer Margin'].isna()]

        # Keep only rows with valid Country AND Country Group (filters out aggregation rows)
        df = df[(df['Country'].notna()) & (df['Country Group'].notna())]

        # Set Customer Margin to 0 for revenue rows
        df['Customer Margin'] = 0

        # Convert all date columns to numeric
        for col in date_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        # Filter to FY27 period (2026-02 to 2027-01)
        df = self.filter_fy27_columns(df)

        return df

    def load_overheads(self) -> pd.DataFrame:
        """Load overhead costs data"""
        df = pd.read_excel(self.file_path, sheet_name='Overheads', header=1)

        # Clean column names
        new_cols = []
        for col in df.columns:
            if isinstance(col, datetime):
                new_cols.append(col.strftime('%Y-%m'))
            else:
                new_cols.append(str(col))
        df.columns = new_cols

        # Keep structure columns and date columns
        base_cols = ['Territory', 'Function', 'Category', 'Group', 'Supplier']
        date_cols = [c for c in df.columns if c.startswith('202')]

        available_cols = [c for c in base_cols if c in df.columns] + date_cols
        df = df[available_cols].copy()
        df = df.dropna(subset=['Territory', 'Category'])

        # Convert all date columns to numeric
        for col in date_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        # Filter to FY27 period (2026-02 to 2027-01)
        df = self.filter_fy27_columns(df)

        return df

    def load_fulfilment_rates(self) -> pd.DataFrame:
        """Load fulfilment cost rates by country and channel"""
        df = pd.read_excel(self.file_path, sheet_name='Fulfilment')
        # Handle column name with trailing space
        rate_col = [c for c in df.columns if 'fulfilment' in c.lower()][0] if any('fulfilment' in c.lower() for c in df.columns) else df.columns[2]
        df = df[['Country', 'Category', rate_col]].copy()
        df = df.rename(columns={'Category': 'Channel', rate_col: 'Rate'})
        df['Rate'] = pd.to_numeric(df['Rate'], errors='coerce').fillna(-0.15)
        return df

    def load_amazon_data(self) -> pd.DataFrame:
        """Load Amazon marketplace data"""
        df = pd.read_excel(self.file_path, sheet_name='Amazon', header=1)

        # Clean column names
        new_cols = []
        for col in df.columns:
            if isinstance(col, datetime):
                new_cols.append(col.strftime('%Y-%m'))
            else:
                new_cols.append(str(col))
        df.columns = new_cols

        # Convert all date columns to numeric
        date_cols = [c for c in df.columns if c.startswith('202')]
        for col in date_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        # Filter to FY27 period (2026-02 to 2027-01)
        df = self.filter_fy27_columns(df)

        return df

    def load_payroll(self) -> pd.DataFrame:
        """Load payroll data"""
        df = pd.read_excel(self.file_path, sheet_name='Payroll', header=4)
        return df

    def load_dtc_inputs(self, territory: str) -> pd.DataFrame:
        """Load DTC input data for a specific territory including subscription cohort data"""
        if territory not in self.territories or territory == 'ROW':
            return pd.DataFrame()

        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb[territory]

            # Extended metrics including subscription cohort data
            data = []
            metric_rows = {
                # Traffic & Conversion
                'Traffic': 4,
                'Conversion Rate': 6,
                'Total Orders': 7,

                # Customer Segmentation
                'New Customers (Non-Subs)': 9,      # Fixed: was 8
                'New Subscription Customers': 10,    # Fixed: was 9
                'Recurring Subscription Customers': 11,  # Fixed: was 10
                'Returning Customers (Non-Subs)': 12,    # Fixed: was 11

                # Revenue by Segment - Returning
                'Returning AOV (Non-Subs)': 15,      # Fixed: was 14
                'Returning Revenue (Non-Subs)': 16,  # Fixed: was 15
                'Returning AOV (Subs)': 17,          # Fixed: was 16
                'Returning Revenue (Subs)': 18,      # Fixed: was 17

                # Revenue by Segment - New
                'New Customer AOV (Non-Subs)': 20,       # Fixed: was 19
                'New Customer Revenue (Non-Subs)': 21,   # Fixed: was 20
                'New Customer AOV (Subs)': 22,           # Fixed: was 21
                'New Customer Revenue (Subs)': 23,       # Fixed: was 22

                # Totals & Adjustments
                'Total Revenue': 25,  # CRITICAL FIX: was 24 (formula cell) - now uses Actual Revenue row
                'Actual Revenue': 25,
                'Missing Revenue (Cohort Adjustment)': 27,  # Fixed: was 26
                'Marketing Budget': 36,  # Row 36 is the actual Marketing Budget total
            }

            # Get month columns (starting from column 5)
            all_months = []
            for col in range(5, min(42, ws.max_column + 1)):  # Extended to capture more months
                val = ws.cell(row=2, column=col).value
                if val and isinstance(val, datetime):
                    month_str = val.strftime('%Y-%m')
                    all_months.append((col, month_str))

            # Filter to FY27 period (2026-02 to 2027-01)
            months = [(col, month) for col, month in all_months if self.fy27_start <= month <= self.fy27_end]

            for metric, row in metric_rows.items():
                row_data = {'Metric': metric, 'Territory': territory}
                for col_num, month in months:
                    val = ws.cell(row=row, column=col_num).value
                    # Handle numeric conversion
                    if val is not None:
                        try:
                            row_data[month] = float(val)
                        except (ValueError, TypeError):
                            row_data[month] = 0
                    else:
                        row_data[month] = 0
                data.append(row_data)

            wb.close()
            return pd.DataFrame(data)
        except Exception as e:
            print(f"Error loading DTC for {territory}: {e}")
            return pd.DataFrame()

    def get_date_columns(self) -> list:
        """Get list of FY27 date columns from the model"""
        df = pd.read_excel(self.file_path, sheet_name='B2B', header=5, nrows=1)
        all_dates = []
        for col in df.columns:
            if isinstance(col, datetime):
                all_dates.append(col.strftime('%Y-%m'))

        # Filter to FY27 period only (2026-02 to 2027-01)
        fy27_dates = [d for d in all_dates if self.fy27_start <= d <= self.fy27_end]
        return fy27_dates

    def load_cogs_rates(self) -> dict:
        """Load CoGS rates from UK P&L (used as defaults)"""
        wb = load_workbook(self.file_path, data_only=False)
        ws = wb['UK P&L']

        rates = {
            'DTC': abs(float(ws.cell(row=16, column=3).value or 0.24)),
            'B2B': abs(float(ws.cell(row=17, column=3).value or 0.26)),
            'Marketplace': abs(float(ws.cell(row=18, column=3).value or 0.18)),
            'TikTok': abs(float(ws.cell(row=19, column=3).value or 0.24)),
        }
        wb.close()
        return rates

    def load_actuals_data(self) -> pd.DataFrame:
        """Load actuals data from Actuals sheet if it exists"""
        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)

            # Check if Actuals sheet exists
            if 'Actuals' not in wb.sheetnames:
                wb.close()
                return pd.DataFrame()

            # Read the Actuals sheet
            df = pd.read_excel(self.file_path, sheet_name='Actuals', header=0)

            # Clean column names - convert datetime to string
            new_cols = []
            for col in df.columns:
                if isinstance(col, datetime):
                    new_cols.append(col.strftime('%Y-%m'))
                else:
                    new_cols.append(str(col))
            df.columns = new_cols

            # Expected columns: Channel, Territory (optional), then date columns
            # Validate structure
            required_cols = ['Channel']
            if not all(c in df.columns for c in required_cols):
                print("Warning: Actuals sheet missing required columns")
                wb.close()
                return pd.DataFrame()

            # Get date columns and convert to numeric
            date_cols = [c for c in df.columns if c.startswith('202')]
            for col in date_cols:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

            # Filter to FY27 period (2026-02 to 2027-01)
            df = self.filter_fy27_columns(df)

            wb.close()
            return df

        except Exception as e:
            print(f"Error loading actuals data: {e}")
            return pd.DataFrame()


def validate_data(data: dict) -> list:
    """Validate loaded data and return list of warnings"""
    warnings = []

    # Check B2B data
    if 'b2b' in data and not data['b2b'].empty:
        b2b = data['b2b']

        # Check for expected columns
        expected_b2b_cols = ['Customer Name', 'Country', 'Country Group']
        missing_b2b = [c for c in expected_b2b_cols if c not in b2b.columns]
        if missing_b2b:
            warnings.append(f"B2B data missing columns: {', '.join(missing_b2b)}")

        # Check for customer margin data
        if 'Customer Margin' in b2b.columns:
            if b2b['Customer Margin'].sum() == 0:
                warnings.append("All B2B Customer Margins are 0 - check if margin data exists in Excel file")

        # Check for revenue data
        date_cols = [c for c in b2b.columns if c.startswith('202')]
        if date_cols:
            total_revenue = b2b[date_cols].sum().sum()
            if total_revenue == 0:
                warnings.append("No B2B revenue data found in date columns")

    # Check overheads data
    if 'overheads' in data and not data['overheads'].empty:
        oh = data['overheads']
        expected_oh_cols = ['Territory', 'Category']
        missing_oh = [c for c in expected_oh_cols if c not in oh.columns]
        if missing_oh:
            warnings.append(f"Overheads data missing columns: {', '.join(missing_oh)}")

    # Check DTC data
    if 'dtc' in data:
        if not data['dtc']:
            warnings.append("No DTC territory data loaded - check if territory sheets exist in Excel")
        else:
            dtc_count = len(data['dtc'])
            if dtc_count < 3:
                warnings.append(f"Only {dtc_count} DTC territories loaded - expected at least 3")

    return warnings


def load_all_data(file_path: str) -> dict:
    """Load all data from the budget model"""
    loader = BudgetDataLoader(file_path)

    data = {
        'b2b': loader.load_b2b_data(),
        'overheads': loader.load_overheads(),
        'fulfilment': loader.load_fulfilment_rates(),
        'amazon': loader.load_amazon_data(),
        'cogs_rates': loader.load_cogs_rates(),
        'actuals': loader.load_actuals_data(),
        'dates': loader.get_date_columns(),
        'territories': loader.territories,
        'territory_groups': loader.territory_groups,
        'channels': loader.channels,
    }

    # Load DTC data for each territory
    data['dtc'] = {}
    for territory in ['UK', 'ES', 'IT', 'RO', 'CZ', 'HU', 'SK', 'Other EU']:
        dtc_data = loader.load_dtc_inputs(territory)
        if not dtc_data.empty:
            data['dtc'][territory] = dtc_data

    # Validate data and store warnings
    data['validation_warnings'] = validate_data(data)

    return data
